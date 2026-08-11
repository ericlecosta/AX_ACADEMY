from pathlib import Path
from openpyxl import Workbook, load_workbook

class PlanilhaMestra:

    def __init__(self, caminho):
        self.caminho = Path(caminho)

        if self.caminho.exists():
            self.wb = load_workbook(self.caminho)
            self.ws = self.wb.active
        else:
            self.wb = Workbook()
            self.ws = self.wb.active

            self.ws.append([
                "Protocolo",
                "Nome",
                "Sobrenome",
                "CPF",
                "Email",
                "Telefone",
                "Endereço",
                "Status"
            ])

            self.wb.save(self.caminho)

    def adicionar_cliente(self, dados):

        self.ws.append([
            dados["protocolo"],
            dados["nome"],
            dados["sobrenome"],
            dados["cpf"],
            dados["email"],
            dados["telefone"],
            dados["endereco"],
            dados["status"]
        ])

        self.wb.save(self.caminho)